from flask import Flask, render_template, request, jsonify, session, redirect, url_for
import os
import json
import requests
import openpyxl
from functools import wraps
import ast
import re

# 秋季周末到日期的映射（可按需调整年份与日期）
AUTUMN_WEEKDAY_TO_DATE = {
    '周六': '2025-09-06',
    '周日': '2025-09-07',
}

def normalize_date_string(date_str: str) -> str:
    """将 20250906 规范为 2025-09-06；已是 YYYY-MM-DD 则原样返回，其它保持原样。"""
    if not isinstance(date_str, str):
        return date_str
    # 纯数字 8 位：YYYYMMDD -> YYYY-MM-DD
    if re.fullmatch(r"\d{8}", date_str):
        return f"{date_str[0:4]}-{date_str[4:6]}-{date_str[6:8]}"
    # 已是 YYYY-MM-DD
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_str):
        return date_str
    return date_str

app = Flask(__name__)
app.secret_key = 'your-secret-key'  # 用于session加密

# 获取应用程序根目录
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 密码文件路径
PASSWORDS_FILE = os.path.join(BASE_DIR, 'passwords.json')

# 初始化密码文件
def init_passwords_file():
    if not os.path.exists(PASSWORDS_FILE):
        with open(PASSWORDS_FILE, 'w') as f:
            json.dump({}, f)

# 读取密码文件
def read_passwords():
    if not os.path.exists(PASSWORDS_FILE):
        return {}
    with open(PASSWORDS_FILE, 'r') as f:
        return json.load(f)

# 保存密码文件
def save_passwords(passwords):
    with open(PASSWORDS_FILE, 'w') as f:
        json.dump(passwords, f)

# 初始化密码文件
init_passwords_file()

# 登录验证装饰器
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/')
def login():
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login_post():
    username = request.form.get('username')
    password = request.form.get('password')
    
    print(f"Login attempt - Username: {username}")
    print(f"Current working directory: {os.getcwd()}")
    print(f"BASE_DIR: {BASE_DIR}")
    
    # 检查用户token文件夹是否存在
    token_dir = os.path.join(BASE_DIR, 'token', username)
    print(f"Checking token directory: {token_dir}")
    
    # 列出token目录下的所有文件夹
    token_base_dir = os.path.join(BASE_DIR, 'token')
    print(f"Available users in token directory:")
    for item in os.listdir(token_base_dir):
        if os.path.isdir(os.path.join(token_base_dir, item)):
            print(f"- {item}")
    
    if not os.path.exists(token_dir):
        print(f"Token directory not found: {token_dir}")
        return jsonify({'success': False, 'error': '用户不存在'})
    
    # 检查必要的文件是否存在
    required_files = ['gateway_app_parameters.xlsx', 'wxbackend_parameters.xlsx', 'wxbackend_sendmsg_parameters.xlsx']
    missing_files = []
    
    for file in required_files:
        file_path = os.path.join(token_dir, file)
        if not os.path.exists(file_path):
            missing_files.append(file)
    
    if missing_files:
        print(f"Missing files in token directory: {missing_files}")
        return jsonify({'success': False, 'error': f'缺少必要的文件: {", ".join(missing_files)}'})
    
    # 验证密码
    passwords = read_passwords()
    user_password = passwords.get(username, '123456')  # 如果用户没有设置密码，使用默认密码
    
    if password != user_password:
        return jsonify({'success': False, 'error': '密码错误'})
    
    session['username'] = username
    return jsonify({'success': True})

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')

def get_staff_token(gateway_app_params_dict):
    # 直接使用参数文件中的accessToken
    access_token = gateway_app_params_dict.get('accessToken')
    if not access_token:
        print("参数文件中没有找到accessToken")
        return None
    return access_token


# 主要修复点：
# 1. 移动filtered_classes初始化到正确位置
# 2. 使用set去重避免重复处理
# 3. 改进筛选逻辑
# 4. 添加调试信息

@app.route('/get_students', methods=['POST'])
@login_required
def get_students():
    try:
        selected_grade = request.json.get('grade', [])
        raw_times = request.json.get('time', [])
        selected_time = []

        # 解析时间数据
        for item in raw_times:
            try:
                parsed = json.loads(item)
                if isinstance(parsed, list):
                    selected_time.extend(parsed)
            except json.JSONDecodeError:
                selected_time.append(item)

        # 秋季适配：将“周六/周日”映射为具体日期，并统一日期格式
        mapped_time = []
        for t in selected_time:
            # 先把周末关键字映射为具体日期
            mapped = AUTUMN_WEEKDAY_TO_DATE.get(t, t)
            # 再做日期格式规范化
            mapped_time.append(normalize_date_string(mapped))
        selected_time = mapped_time

        # 关键调试：检查接收到的参数
        print("=" * 50)
        print(f"🔍 调试信息 - 接收到的参数:")
        print(f"原始年级数据: {request.json.get('grade')} (类型: {type(request.json.get('grade'))})")
        print(f"原始时间数据: {request.json.get('time')} (类型: {type(request.json.get('time'))})")
        print(f"处理后的年级: {selected_grade}")
        print(f"处理后的时间: {selected_time}")
        print("=" * 50)

        # 参数验证 - 添加更严格的验证
        if not selected_grade:
            print("❌ 错误: 未选择年级")
            return jsonify({'success': False, 'error': '请选择年级'})

        if not selected_time:
            print("❌ 错误: 未选择时间")
            return jsonify({'success': False, 'error': '请选择时间'})

        # 获取参数文件
        username = session['username']
        token_dir = os.path.join(BASE_DIR, 'token', username)
        gateway_params_file = os.path.join(token_dir, 'gateway_app_parameters.xlsx')
        wxbackend_params_file = os.path.join(token_dir, 'wxbackend_parameters.xlsx')

        # 读取参数
        gateway_app_params_wb = openpyxl.load_workbook(gateway_params_file)
        wxbackend_params_wb = openpyxl.load_workbook(wxbackend_params_file)

        gateway_app_params_dict = {}
        for row in gateway_app_params_wb.active.iter_rows(min_row=2, values_only=True):
            parameter = row[0]
            value = row[1] if row[1] is not None else ''
            gateway_app_params_dict[parameter] = value

        wxbackend_params_dict = {}
        for row in wxbackend_params_wb.active.iter_rows(min_row=2, values_only=True):
            parameter = row[0]
            value = row[1] if row[1] is not None else ''
            wxbackend_params_dict[parameter] = value

        # 获取accessToken
        headers_get_token = {
            'Origin': 'https://deskwx.xdf.cn',
            'Connection': 'keep-alive',
            'Accept': 'application/json, text/plain, */*',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Language/zh ColorScheme/Light DistType/publish-store wxwork/4.1.15 (MicroMessenger/6.2) WeChat/2.0.4 Safari/605.1.15',
            'Accept-Language': 'zh-cn',
            'Referer': 'https://deskwx.xdf.cn/',
        }

        params_get_token = {
            'appId': '',
            'appVersion': '',
            'accessToken': '',
        }

        max_retries = 3
        retries = 0
        while retries < max_retries:
            try:
                response_get_token = requests.get('https://wxbackend.xdf.cn/api/wx/getToken',
                                                  params=params_get_token,
                                                  headers=headers_get_token)
                break
            except requests.exceptions.RequestException as e:
                print(f"Connection error: {e}")
                retries += 1

        access_token = response_get_token.json().get('data', {}).get('accessToken', '')
        print(f"🔑 获取到的访问令牌: {access_token[:10]}...")

        # 设置请求参数
        headers = {
            'Host': 'gateway.app.xdf.cn',
            'stafftoken': '98ba3907b83fd78eba5463580010e28c',
            'accept': 'application/json, text/plain, */*',
            'origin': 'https://deskwx.xdf.cn',
            'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Language/zh ColorScheme/Light DistType/publish-store wxwork/4.1.15 (MicroMessenger/6.2) WeChat/2.0.4 Safari/605.1.15',
            'accept-language': 'zh-cn',
            'referer': 'https://deskwx.xdf.cn/wechat-work-teacher-ms-web/classSchedule/classList',
        }

        params = {
            'appId': '3027D9E5-0C09-4AD3-86F0-6C678B7826A4',
            'appVersion': '',
            'accessToken': access_token,
            'userId': gateway_app_params_dict.get('userId', ''),
            'classStatus': '3',
            'schoolId': '3',
            'teacherCode': gateway_app_params_dict.get('teacherCode', ''),
            'teacherType': '1',
            'pageSize': '5',
            'pageNo': '1',
        }

        # 更新参数
        for key, value in gateway_app_params_dict.items():
            if key != 'Host' and key != 'accessToken' and key != 'pageNo':
                if key in headers:
                    headers[key] = value
                if key in params:
                    params[key] = value

        # 收集所有班级数据
        all_classes = []
        print(f"📚 开始获取班级列表...")

        for index in range(1, 4):
            params['pageNo'] = index
            print(f"  📄 获取第 {index} 页班级数据...")

            response = requests.get(
                'https://gateway.app.xdf.cn/k12-assistant-api/api/v1.0/acl/wx/1/class/all/list',
                params=params,
                headers=headers,
            )

            data = response.json()
            if 'data' in data and 'classList' in data['data']:
                page_classes = data['data']['classList']
                all_classes.extend(page_classes)
                print(f"    ✅ 第 {index} 页获取到 {len(page_classes)} 个班级")
            else:
                print(f"    ❌ 第 {index} 页获取失败: {data}")

        print(f"📊 总共获取到 {len(all_classes)} 个班级")

        # 调试：显示所有班级的基本信息
        print("\n🔍 所有班级详情:")
        for i, cls in enumerate(all_classes[:10]):  # 只显示前10个避免输出过多
            print(f"  {i + 1}. 班级名: {cls.get('className', 'N/A')}")
            print(f"     开课日期: {cls.get('classStartDate', 'N/A')}")
            print(f"     班级代码: {cls.get('classCode', 'N/A')}")
            print(f"     包含'暑假': {'暑假' in cls.get('className', '')}")
            print("")

        # 筛选符合条件的班级
        print(f"🎯 开始筛选班级...")
        print(f"筛选条件:")
        print(f"  - 必须包含'暑假'")
        print(f"  - 年级匹配: {selected_grade}")
        print(f"  - 时间匹配: {selected_time}")
        print("")

        filtered_classes = []
        processed_classes = set()

        for class_item in all_classes:
            class_name = class_item.get('className', '')
            class_start_date = class_item.get('classStartDate', '')
            class_code = class_item.get('classCode', '')

            # 避免重复处理
            if class_code in processed_classes:
                continue

            print(f"🔍 检查班级: {class_name}")
            print(f"   开课日期: {class_start_date}")

            # 1. 检查是否包含'暑假'
            has_summer = '暑假' in class_name
            print(f"   包含'暑假': {has_summer}")
            if not has_summer:
                print("   ❌ 不包含'暑假'，跳过")
                continue

            # 2. 检查年级匹配 - 改进匹配逻辑
            grade_match = False
            matched_grade = None
            for grade in selected_grade:
                # 确保年级匹配是精确的，避免"一年级"匹配到"十一年级"
                if grade in class_name:
                    # 进一步验证是否为精确匹配
                    # 可以根据实际班级名称格式调整这个逻辑
                    grade_match = True
                    matched_grade = grade
                    break

            print(f"   年级匹配: {grade_match} (匹配的年级: {matched_grade})")
            if not grade_match:
                print("   ❌ 年级不匹配，跳过")
                continue

            # 3. 检查时间匹配 - 改进时间匹配逻辑
            time_match = False
            matched_time = None
            for time in selected_time:
                # 精确匹配开课日期
                if time == class_start_date:
                    time_match = True
                    matched_time = time
                    break
                # 如果日期格式不完全一致，也尝试包含匹配
                elif time in class_start_date:
                    time_match = True
                    matched_time = time
                    break

            print(f"   时间匹配: {time_match} (匹配的时间: {matched_time})")
            if not time_match:
                print("   ❌ 时间不匹配，跳过")
                continue

            # 如果所有条件都匹配，添加到结果
            filtered_classes.append((class_code, class_name))
            processed_classes.add(class_code)
            print(f"   ✅ 符合条件！添加到结果列表")
            print("")

        print(f"🎉 筛选完成！符合条件的班级数量: {len(filtered_classes)}")

        if not filtered_classes:
            print("⚠️  没有找到符合条件的班级")
            return jsonify({
                'success': True,
                'students': [],
                'message': '没有找到符合条件的班级'
            })

        # 获取每个班级的学生列表 - 修复版本
        unfriend_students = []

        for class_code, class_name in filtered_classes:
            print(f"\n👥 处理班级学生: {class_name} (班级代码: {class_code})")

            headers_1 = {
                'Host': 'wxbackend.xdf.cn',
                'accept': 'application/json, text/plain, */*',
                'origin': 'https://deskwx.xdf.cn',
                'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Language/zh ColorScheme/Light DistType/publish-store wxwork/4.1.15 (MicroMessenger/6.2) WeChat/2.0.4 Safari/605.1.15',
                'accept-language': 'zh-cn',
                'referer': 'https://deskwx.xdf.cn/wechat-work-teacher-ms-web/classSchedule/classList',
            }

            params_1 = {
                'appId': '3027D9E5-0C09-4AD3-86F0-6C678B7826A4',
                'appVersion': '',
                'accessToken': access_token,
                'userId': gateway_app_params_dict.get('userId', ''),
                'classCode': class_code,  # 这是关键参数，不应该被覆盖
                'schoolId': '3',
                'classGroupCode': '',
                'chatId': '',
            }

            # 更新参数 - 修复版本：排除会覆盖关键参数的字段
            critical_params = {'classCode', 'accessToken', 'Host'}  # 不应该被覆盖的关键参数

            for key, value in wxbackend_params_dict.items():
                if key not in critical_params:  # 只更新非关键参数
                    if key in headers_1:
                        headers_1[key] = value
                    if key in params_1:
                        params_1[key] = value

            # 🔍 调试：打印请求参数
            print(f"   📋 请求参数:")
            print(f"     classCode: {params_1['classCode']}")
            print(f"     userId: {params_1['userId']}")
            print(f"     accessToken: {params_1['accessToken'][:10]}...")

            # 🔍 额外调试：确认参数没有被意外覆盖
            if params_1['classCode'] != class_code:
                print(f"   🚨 警告：classCode参数被意外覆盖！")
                print(f"     期望值: {class_code}")
                print(f"     实际值: {params_1['classCode']}")
                # 强制恢复正确的值
                params_1['classCode'] = class_code
                print(f"     已强制恢复为: {params_1['classCode']}")

            # 添加随机延迟避免API限制
            import time
            import random
            time.sleep(random.uniform(0.1, 0.3))

            try:
                # 获取学生列表
                response_1 = requests.get(
                    'https://wxbackend.xdf.cn/api/wx/class/getStudentList',
                    params=params_1,
                    headers=headers_1,
                    timeout=10
                )

                # 🔍 调试：打印完整的API响应
                print(f"   📡 API响应状态码: {response_1.status_code}")
                print(f"   📡 API请求URL: {response_1.url}")

                # 验证URL中的classCode是否正确
                if f"classCode={class_code}" not in response_1.url:
                    print(f"   🚨 严重错误：URL中的classCode不匹配！")
                    print(f"   🔍 期望在URL中看到: classCode={class_code}")
                    print(f"   🔍 实际URL: {response_1.url}")

                data_1 = response_1.json()

                # 🔍 调试：打印响应数据状态
                print(f"   📊 响应数据状态: {data_1.get('status')}")
                if 'data' in data_1 and data_1['data']:
                    if 'studentList' in data_1['data']:
                        student_count = len(data_1['data']['studentList'])
                        print(f"   📊 API返回学生数量: {student_count}")

                        # 打印前几个学生的姓名用于调试
                        if student_count > 0:
                            first_few_students = data_1['data']['studentList'][:3]
                            student_names = [s.get('studentName', 'N/A') for s in first_few_students]
                            print(f"   👤 前几个学生: {student_names}")
                    else:
                        print("   ❌ 响应中没有studentList字段")
                else:
                    print("   ❌ 响应中没有data字段或data为空")
                    print(f"   🔍 完整响应: {data_1}")

                if data_1.get('status') != 100000:
                    print(f"   ❌ 获取学生列表失败: {data_1.get('message')}")
                    continue

                if data_1.get('data') is None:
                    print(f"   ⚠️  班级 {class_name} 返回空数据")
                    continue

                if data_1['data'].get('studentList') is None:
                    print(f"   ⚠️  班级 {class_name} 没有学生列表")
                    continue

                # 统计学生信息
                all_students = data_1['data']['studentList']
                class_unfriend_students = []

                # 🔍 调试：检查学生数据
                print(f"   🔍 开始处理 {len(all_students)} 个学生...")

                for i, student in enumerate(all_students):
                    student_name = student.get('studentName', f'学生{i + 1}')
                    student_code = student.get('studentCode', '')
                    friend_status = student.get('friend', 0)

                    # 只在前5个学生时打印详细信息
                    if i < 5:
                        print(f"     学生{i + 1}: {student_name} (好友状态: {friend_status})")

                    if friend_status != 1:
                        student_info = {
                            'name': student_name,
                            'code': student_code,
                            'classCode': class_code,
                            'className': class_name,
                            'friendStatus': friend_status  # 添加好友状态用于调试
                        }
                        class_unfriend_students.append(student_info)

                # 检查是否有重复的学生代码
                student_codes = [s['code'] for s in class_unfriend_students]
                unique_codes = set(student_codes)
                if len(student_codes) != len(unique_codes):
                    print(f"   ⚠️  发现重复的学生代码！")

                unfriend_students.extend(class_unfriend_students)

                print(f"   📊 班级 {class_name} 学生统计:")
                print(f"     总学生数: {len(all_students)}")
                print(f"     未添加好友: {len(class_unfriend_students)}")
                print(f"     已添加好友: {len(all_students) - len(class_unfriend_students)}")

                # 🔍 如果这个班级的学生数和之前的完全一样，特别标注
                if len(all_students) == 33 and len(class_unfriend_students) == 29:
                    print(f"   🚨 警告：这个班级的数据和之前的班级完全一样！")
                    print(f"   🔍 可能的问题：API忽略了classCode参数或返回了缓存数据")

            except requests.exceptions.RequestException as e:
                print(f"   ❌ 网络请求失败: {e}")
                continue
            except Exception as e:
                print(f"   ❌ 处理班级 {class_name} 时出错: {e}")
                continue

        # 🔍 最终数据分析
        print(f"\n🎯 最终数据分析:")
        print(f"   总学生数: {len(unfriend_students)}")

        # 按班级分组统计
        from collections import defaultdict
        class_stats = defaultdict(int)
        for student in unfriend_students:
            class_stats[student['className']] += 1

        print(f"   按班级统计:")
        for class_name, count in class_stats.items():
            print(f"     {class_name}: {count}人")

        # 检查是否有重复学生
        all_student_codes = [s['code'] for s in unfriend_students]
        unique_student_codes = set(all_student_codes)
        if len(all_student_codes) != len(unique_student_codes):
            print(f"   ⚠️  发现 {len(all_student_codes) - len(unique_student_codes)} 个重复学生")

        # 去重处理
        unique_students = []
        seen_students = set()
        for student in unfriend_students:
            student_key = (student['code'], student['classCode'])
            if student_key not in seen_students:
                unique_students.append(student)
                seen_students.add(student_key)

        print(f"\n🎯 最终结果:")
        print(f"   处理的班级数: {len(filtered_classes)}")
        print(f"   未添加好友的学生总数: {len(unique_students)}")
        print("=" * 50)

        return jsonify({
            'success': True,
            'students': unique_students,
            'debug_info': {
                'total_classes_found': len(all_classes),
                'filtered_classes_count': len(filtered_classes),
                'filtered_classes': [name for _, name in filtered_classes],
                'selected_grade': selected_grade,
                'selected_time': selected_time
            }
        })

    except Exception as e:
        print(f"❌ Error in get_students: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get_username')
@login_required
def get_username():
    return jsonify({
        'success': True,
        'username': session.get('username', '')
    })

@app.route('/get_default_message')
@login_required
def get_default_message():
    try:
        username = session['username']
        token_dir = os.path.join(BASE_DIR, 'token', username)
        params_file = os.path.join(token_dir, 'wxbackend_sendmsg_parameters.xlsx')
        
        wb = openpyxl.load_workbook(params_file)
        ws = wb.active
        
        # 查找content参数
        content = ''
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == 'content':
                content = row[1] if row[1] is not None else ''
                break
        
        return jsonify({
            'success': True,
            'message': content
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })

@app.route('/send_message', methods=['POST'])
@login_required
def send_message():
    try:
        message = request.json.get('message')
        selected_students = request.json.get('students')
        
        print("=" * 50)
        print("接收到的原始数据：")
        print(f"request.json: {request.json}")
        print(f"message: {message}")
        print(f"selected_students: {selected_students}")
        print(f"selected_students类型: {type(selected_students)}")
        if selected_students and len(selected_students) > 0:
            print(f"第一个学生数据: {selected_students[0]}")
            print(f"第一个学生数据类型: {type(selected_students[0])}")
        print("=" * 50)
        
        if not message:
            return jsonify({'success': False, 'error': '消息内容不能为空'})
            
        if not selected_students or not isinstance(selected_students, list) or len(selected_students) == 0:
            return jsonify({'success': False, 'error': '请选择至少一个学生'})
        
        username = session['username']
        token_dir = os.path.join(BASE_DIR, 'token', username)
        
        # 读取参数文件
        gateway_params_file = os.path.join(token_dir, 'gateway_app_parameters.xlsx')
        wxbackend_params_file = os.path.join(token_dir, 'wxbackend_parameters.xlsx')
        wxbackendmsg_params_file = os.path.join(token_dir, 'wxbackend_sendmsg_parameters.xlsx')
        
        # 读取参数
        gateway_app_params_wb = openpyxl.load_workbook(gateway_params_file)
        wxbackend_params_wb = openpyxl.load_workbook(wxbackend_params_file)
        wxbackendmsg_params_wb = openpyxl.load_workbook(wxbackendmsg_params_file)
        
        gateway_app_params_dict = {}
        for row in gateway_app_params_wb.active.iter_rows(min_row=2, values_only=True):
            parameter = row[0]
            value = row[1] if row[1] is not None else ''
            gateway_app_params_dict[parameter] = value
            
        wxbackend_params_dict = {}
        for row in wxbackend_params_wb.active.iter_rows(min_row=2, values_only=True):
            parameter = row[0]
            value = row[1] if row[1] is not None else ''
            wxbackend_params_dict[parameter] = value

        wxbackendmsg_params_dict = {}
        for row in wxbackendmsg_params_wb.active.iter_rows(min_row=2, values_only=True):
            parameter = row[0]
            value = row[1] if row[1] is not None else ''
            wxbackendmsg_params_dict[parameter] = value
        
        # 获取accessToken
        headers_get_token = {
            'Origin': 'https://deskwx.xdf.cn',
            'Connection': 'keep-alive',
            'Accept': 'application/json, text/plain, */*',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Language/zh ColorScheme/Light DistType/publish-store wxwork/4.1.15 (MicroMessenger/6.2) WeChat/2.0.4 Safari/605.1.15',
            'Accept-Language': 'zh-cn',
            'Referer': 'https://deskwx.xdf.cn/',
        }
        
        params_get_token = {
            'appId': '',
            'appVersion': '',
            'accessToken': '',
        }
        
        response_get_token = requests.get('https://wxbackend.xdf.cn/api/wx/getToken',
                                        params=params_get_token,
                                        headers=headers_get_token)
        
        access_token = response_get_token.json().get('data', {}).get('accessToken', '')
        
        # 构建请求数据
        final_json_data = {
            'appId': '3027D9E5-0C09-4AD3-86F0-6C678B7826A4',
            'appVersion': '',
            'accessToken': access_token,
            'wxUserId': '',
            'classReq': {
                'schoolId': '3',
                'classCode': '',
                'classGroupCode': '',
                'classType': '1'
            },
            'content': message,
            'teacher': None,
            'studentList': []  # 初始化为空列表
        }

        # 从参数文件中更新值
        print("=" * 50)
        print("开始处理参数文件中的值：")
        for key, value in wxbackendmsg_params_dict.items():
            # 跳过studentList，因为它应该从selected_students获取
            if key == 'studentList':
                continue
            # 跳过content，保持前端传入的message
            if key == 'content':
                continue
            if key != 'Host' and key != 'accessToken' and key != 'teacher' and key != 'classReq':
                if key in final_json_data:
                    final_json_data[key] = value
                    print(f"设置 {key} = {value}")
            elif key == 'teacher':
                if key in final_json_data:
                    try:
                        teacher_dict = ast.literal_eval(value)
                        final_json_data[key] = teacher_dict
                        print(f"设置 teacher = {teacher_dict}")
                    except Exception as e:
                        print(f"解析teacher失败: {e}")
                        print(f"原始值: {value}")
                        final_json_data[key] = None
            elif key == 'classReq':
                if key in final_json_data:
                    try:
                        # 解析classReq字符串为字典
                        class_req_str = value.replace("'", '"')
                        class_req_dict = json.loads(class_req_str)
                        final_json_data[key] = class_req_dict
                        print(f"设置 classReq = {class_req_dict}")
                    except Exception as e:
                        print(f"解析classReq失败: {e}")
                        print(f"原始值: {value}")
                        # 保持默认值
                        final_json_data[key] = {
                            'schoolId': '3',
                            'classCode': '',
                            'classGroupCode': '',
                            'classType': '1'
                        }

        # 处理学生列表
        print("处理学生列表：")
        final_json_data['studentList'] = selected_students  # 直接使用selected_students
        print(f"设置 studentList = {selected_students}")

        print("=" * 50)
        print("最终构造的数据：")
        print(json.dumps(final_json_data, indent=2, ensure_ascii=False))
        print("=" * 50)

        # 检查所有必需参数
        print("=" * 50)
        print("参数检查：")
        print(f"1. appId: {final_json_data['appId']}")
        print(f"2. accessToken: {final_json_data['accessToken'][:20]}...")
        print(f"3. wxUserId: {final_json_data['wxUserId']}")
        print(f"4. classReq:")
        print(f"   - schoolId: {final_json_data['classReq']['schoolId']}")
        print(f"   - classCode: {final_json_data['classReq']['classCode']}")
        print(f"   - classGroupCode: {final_json_data['classReq']['classGroupCode']}")
        print(f"   - classType: {final_json_data['classReq']['classType']}")
        print(f"5. content: {final_json_data['content'][:50]}...")
        print(f"6. studentList: {len(final_json_data['studentList'])} 个学生")
        if final_json_data['studentList']:
            print(f"   第一个学生信息:")
            print(f"   - studentCode: {final_json_data['studentList'][0].get('studentCode')}")
            print(f"   - studentName: {final_json_data['studentList'][0].get('studentName')}")
            print(f"   - schoolId: {final_json_data['studentList'][0].get('schoolId')}")
        print("=" * 50)

        # 检查参数文件中的值
        print("参数文件中的值：")
        print(f"wxbackend_params_dict: {wxbackend_params_dict}")
        print(f"wxbackendmsg_params_dict: {wxbackendmsg_params_dict}")
        print("=" * 50)

        print("发送的数据：", final_json_data)
        
        # 发送消息
        response = requests.post(
            'https://wxbackend.xdf.cn/api/wx/sms/send',
            headers=headers_get_token,
            json=final_json_data
        )
        
        result = response.json()
        print("API响应状态码：", response.status_code)
        print("API响应头：", response.headers)
        print("API响应内容：", result)
        
        # 如果发送成功，更新默认消息内容
        if result.get('status') == 100000:
            # 更新wxbackend_sendmsg_parameters.xlsx中的content
            params_file = os.path.join(token_dir, 'wxbackend_sendmsg_parameters.xlsx')
            wb = openpyxl.load_workbook(params_file)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2):
                if row[0].value == 'content':
                    row[1].value = message
                    break
            
            wb.save(params_file)
            
            return jsonify({'success': True})
        else:
            return jsonify({
                'success': False,
                'error': result.get('msg', '发送失败')
            })
            
    except Exception as e:
        print("错误：", str(e))
        return jsonify({'success': False, 'error': str(e)})

@app.route('/change_password', methods=['GET'])
@login_required
def change_password_page():
    return render_template('change_password.html')

@app.route('/change_password', methods=['POST'])
@login_required
def change_password():
    try:
        old_password = request.json.get('oldPassword')
        new_password = request.json.get('newPassword')
        confirm_password = request.json.get('confirmPassword')
        
        if not old_password or not new_password or not confirm_password:
            return jsonify({'success': False, 'error': '所有字段都必须填写'})
            
        if new_password != confirm_password:
            return jsonify({'success': False, 'error': '新密码和确认密码不匹配'})
            
        username = session['username']
        passwords = read_passwords()
        current_password = passwords.get(username, '123456')
        
        if old_password != current_password:
            return jsonify({'success': False, 'error': '当前密码错误'})
            
        # 更新密码
        passwords[username] = new_password
        save_passwords(passwords)
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

if __name__ == '__main__':
    app.run(debug=True)